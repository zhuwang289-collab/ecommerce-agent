"""主流程 — 爬虫 → 优化 → 抖店上架 → 展示 → Excel 导出"""

import sys
from pathlib import Path

from dotenv import load_dotenv
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich import box

load_dotenv()

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.scraper import run as run_scraper
from src.optimizer import run as run_optimizer
from src.douyin_uploader import run as run_uploader

DATA_DIR = PROJECT_ROOT / "data"
console = Console()


# ── 展示 ─────────────────────────────────────────────────

def _fmt_price(price: float, currency: str = "USD") -> str:
    sym = {"USD": "$", "GBP": "£", "EUR": "€", "CNY": "¥"}
    return f"{sym.get(currency, currency)} {price:.2f}"


def display_results(products: list[dict],
                    upload_status: list[str] | None = None) -> None:
    """用 rich 表格打印优化 + 上架状态摘要。"""
    table = Table(
        title="📊 优化 & 上架结果摘要（前 20 条）",
        box=box.ROUNDED,
        header_style="bold cyan",
        border_style="blue",
    )
    table.add_column("#", style="dim", width=3)
    table.add_column("原标题", width=26, overflow="fold")
    table.add_column("品类", width=14, overflow="fold")
    table.add_column("优化标题", width=28, overflow="fold")
    table.add_column("原价", justify="right", width=9)
    table.add_column("建议售价", justify="right", width=9)
    table.add_column("上架状态", justify="center", width=10)

    for i, p in enumerate(products[:20], 1):
        status = upload_status[i - 1] if upload_status else "—"
        status_style = "bold green" if status == "成功" else "bold red"
        currency = p.get("currency", "USD")
        table.add_row(
            str(i),
            p["title"],
            p.get("category", ""),
            p.get("optimized_title", p["title"])[:50],
            _fmt_price(p["price"], currency),
            _fmt_price(p["suggested_price"], currency),
            f"[{status_style}]{status}[/]",
        )

    console.print(table)
    if len(products) > 20:
        total = len(products)
        ok = upload_status.count("成功") if upload_status else 0
        fail = total - ok
        status_line = (f" | ✅ 成功 {ok}  ❌ 失败 {fail}"
                       if upload_status else "")
        console.print(f"  ... 共 {total} 条{status_line}，完整数据见 Excel / JSON")


# ── Excel 导出 ───────────────────────────────────────────

def export_to_excel(products: list[dict],
                    upload_status: list[str] | None = None) -> Path:
    """导出 Excel，含所有字段 + 上架状态 + 价格变动列。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "优化结果"

    columns = [
        ("title", "标题"),
        ("price", "原价"),
        ("currency", "货币"),
        ("suggested_price", "建议售价"),
        ("价格变动", "价格变动"),
        ("optimized_title", "优化标题"),
        ("category", "类别"),
        ("stock_status", "库存状态"),
        ("upc", "UPC"),
        ("上架状态", "上架状态"),
        ("description", "描述"),
        ("image_url", "图片链接"),
        ("detail_url", "商品链接"),
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, (_, label) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, prod in enumerate(products, 2):
        for col_idx, (key, _) in enumerate(columns, 1):
            if key == "价格变动":
                value = 0
            elif key == "上架状态":
                value = (upload_status[row_idx - 2]
                         if upload_status else "—")
            else:
                value = prod.get(key, "")
                if key in ("price", "suggested_price") and isinstance(value, (int, float)):
                    value = round(value, 2)
                if key == "description":
                    value = str(value)[:200] if value else ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    col_widths = [50, 10, 8, 12, 12, 50, 15, 14, 18, 12, 60, 50, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    out_path = DATA_DIR / "optimized_products.xlsx"
    wb.save(out_path)
    return out_path


# ── 主流程 ───────────────────────────────────────────────

def main() -> None:
    console.print(Panel.fit("🛒  E-commerce Automation Agent", border_style="bold green"))

    # ---------- Phase 1: 爬虫 ----------
    console.print("\n[bold yellow]▶ Phase 1: 商品获取（FakeStore API）[/]")
    try:
        products = run_scraper()
    except Exception as e:
        console.print(f"[bold red]✗ 爬取失败: {e}[/]")
        sys.exit(1)

    # ---------- Phase 2: DeepSeek 优化 ----------
    console.print("\n[bold yellow]▶ Phase 2: DeepSeek 标题 & 定价优化[/]")
    try:
        optimized = run_optimizer()
    except Exception as e:
        console.print(f"[bold red]✗ 优化失败: {e}[/]")
        sys.exit(1)

    # ---------- Phase 3: 抖店上架 ----------
    console.print("\n[bold yellow]▶ Phase 3: 抖店商品上架[/]")
    try:
        upload_results = run_uploader(optimized)
        upload_status = [r.status for r in upload_results]
    except Exception as e:
        console.print(f"[red]✗ 上架失败: {e}[/]")
        upload_status = None

    # ---------- Phase 4: 结果展示 ----------
    console.print("\n[bold yellow]▶ Phase 4: 结果展示[/]")
    display_results(optimized, upload_status)

    # ---------- Phase 5: Excel 导出 ----------
    console.print("\n[bold yellow]▶ Phase 5: Excel 导出[/]")
    try:
        xlsx_path = export_to_excel(optimized, upload_status)
        console.print(f"  ✓ 已导出 [bold]{xlsx_path}[/] ({len(optimized)} 条)")
    except Exception as e:
        console.print(f"[red]✗ Excel 导出失败: {e}[/]")

    console.print("\n[bold green]✔ 全部流程执行完毕！[/]")


if __name__ == "__main__":
    main()
